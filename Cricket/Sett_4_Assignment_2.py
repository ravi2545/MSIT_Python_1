import openpyxl
import statistics

wb = openpyxl.load_workbook("cricket.xlsx")
new_wb = openpyxl.Workbook()
new_wb1 = openpyxl.Workbook()
new_sheet = new_wb["Sheet"]
new_wb.remove(new_sheet)
new_sheet1 = new_wb1["Sheet"]
new_wb1.remove(new_sheet1)
sheet_names = wb.sheetnames


for sheet in sheet_names:
    sheet_name= wb[sheet]
    max_r = sheet_name.max_row
    max_c = sheet_name.max_column
    strike_rates = []
    
    for row_line in range(1, max_r+1):
        runs = sheet_name.cell(row = row_line, column = 1).value
        balls = sheet_name.cell(row = row_line, column = 2).value
        strike_rate = runs/balls*100
        strike_rates.append(strike_rate)
    sdv = statistics.stdev(strike_rates)
    new_sheet = new_wb.create_sheet(sheet)
    new_sheet1 = new_wb1.create_sheet(sheet)
    count1 = 1
    count2 = 1
    for line in strike_rates:
        if line<=sdv:
            new_sheet.cell(row=count1, column=1).value=line
            count1+=1
        else:
            new_sheet1.cell(row=count2, column=1).value=line
            count2+=1

new_wb.save("BelowSTD.xlsx")
new_wb1.save("AboveSTD.xlsx")


