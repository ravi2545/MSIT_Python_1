import re
import openpyxl
wb = openpyxl.load_workbook("ElectionsData.xlsx")
sheet_names = wb.sheetnames

for sheet in sheet_names:
    sheet_name = wb[sheet]
    max_r = sheet_name.max_row
    max_c = sheet_name.max_column
    w_count = 0
    total_count = 0
    winner = {}
    for i in range(1, max_r+1):
        winner_female = sheet_name.cell(row = i,column = 4).value
        l_female = sheet_name.cell(row = i, column = 8).value
        
        if 'F' in winner_female:
            total_count = total_count+1
            w_count+=1
        if 'F' in l_female:
            total_count+=1
        winner[sheet]={"Total_Female":total_count, "Winner_Female":w_count}
    print(winner)
