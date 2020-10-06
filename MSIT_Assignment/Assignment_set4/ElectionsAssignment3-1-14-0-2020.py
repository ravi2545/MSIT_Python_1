import openpyxl
wb = openpyxl.load_workbook("ElectionsData.xlsx")
sheet_names = wb.sheetnames
parties = {}
for sheet in sheet_names:
    sheet_name = wb[sheet]
    max_r = sheet_name.max_row
    max_c = sheet_name.max_column
    winner_list = []
    for i in range(1, max_r+1):
        winner_name = sheet_name.cell(row = i, column = 3).value
        winner_votes = sheet_name.cell(row = i, column = 6).value
        l_votes = sheet_name.cell(row = i, column = 10).value
        margin_winner = int(winner_votes)-int(l_votes)
        winner_list.append((margin_winner, winner_name))
    winner_list.sort()
    parties[margin_winner] = winner_list[-1][-1]

print(parties)
