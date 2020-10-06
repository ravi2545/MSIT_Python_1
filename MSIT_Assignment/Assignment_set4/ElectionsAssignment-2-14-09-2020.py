import openpyxl
import re
file_names = ["Elections2004.txt", "Elections2009.txt", "Elections2014.txt"]
wb = openpyxl.Workbook()
sheet = wb["Sheet"]
wb.remove(sheet)
for names in file_names:
    name = re.sub("[A-Za-z.]","",names)
    f_obj = open(names, 'r')
    sheet = wb.create_sheet(name)
    for nline, line in enumerate(f_obj):
        feilds = line.split()
        max_c = 0
        for d in feilds:
            sheet.cell(row=nline+1, column = max_c+1 ).value = d
            max_c+=1

wb.save("ElectionsData.xlsx")
            
