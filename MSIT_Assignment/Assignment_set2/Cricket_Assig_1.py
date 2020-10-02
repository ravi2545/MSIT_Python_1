import openpyxl
import statistics
import re
wb = openpyxl.load_workbook("cricket.xlsx")
sheet_names = wb.sheetnames
sd = {}
for sheet in sheet_names:
    sheet_name = wb[sheet]
    max_r = sheet_name.max_row
    max_c = sheet_name.max_column
    strikerates = []
    for i in range(1, max_r+1):
        runs = sheet_name.cell(row = i,column=1).value
        balls = sheet_name.cell(row = i, column=2).value
        strike_rate = runs/balls*100
        strikerates.append(strike_rate)
    sdv = statistics.stdev(strikerates)
    sd[sheet]=sdv

print(sd)

