import openpyxl
import re
filenames = ["dravid.txt", "sachin.txt", "virat.txt", "yuvraj.txt"]
wb = openpyxl.Workbook()
sheet = wb["Sheet"]
wb.remove(sheet)
for names in filenames:
    name = re.sub("[.txt]","", names)
    f_obj = open(names, 'r')
    sheet = wb.create_sheet(name)
    for linun, data in enumerate(f_obj):
        feilds = data.split()
        runs = int(feilds[2])
        balls = int(feilds[3])
        sheet.cell(row=linun+1, column=1).value = runs
        sheet.cell(row=linun+1, column=2).value = balls

wb.save("cricket.xlsx")
