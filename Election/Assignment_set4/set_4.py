import openpyxl
wb = openpyxl.load_workbook("ElectionsData.xlsx")
wbo = openpyxl.Workbook()
new_sheet = wbo["Sheet"]
wbo.remove(new_sheet)
sheet_names = wb.sheetnames
for sheet in sheet_names:
    sheet_name = wb[sheet]
    max_r = sheet_name.max_row
    max_c = sheet_name.max_column
    new_sheet = wbo.create_sheet(sheet)
    for row_line in range(1, max_r+1):
        winner = sheet_name.cell(row = row_line, column = 3).value
        runner = sheet_name.cell(row = row_line, column = 7).value
        winner_votes = sheet_name.cell(row = row_line, column = 6).value
        l_votes = sheet_name.cell(row = row_line, column = 10).value
        margin_winner = int(winner_votes)-int(l_votes)
        new_sheet.cell(row = row_line, column = 1).value = winner
        new_sheet.cell(row = row_line, column = 2).value = runner
        new_sheet.cell(row = row_line, column = 3).value = margin_winner
wbo.save("newexceldata.xlsx")
        
